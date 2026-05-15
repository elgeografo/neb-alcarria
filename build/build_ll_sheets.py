"""
Genera 7 pestañas exploratorias (LL0..LL6) + pestaña RESUMEN agregando costes.

Estructura idéntica en cada LL:
  - Sección 1: DATOS GENERALES (días, ediciones, asistentes, periodo, ventana)
  - Sección 2: PERSONAL (apoyo, docente, coordinación) con filas GP1 y GP5
    visibles en cada subsección, aunque estén a 0. Inputs en azul.
  - Sección 3: RESUMEN (horas y costes UPM/COPR + total).
  - Fila al final: B.1 Viajes y dietas (referencia a hoja VIAJES).

La pestaña RESUMEN agrega los costes de las 7 pestañas LL en una tabla única.

Uso:
  python3 build/build_ll_sheets.py
ATENCIÓN: cada ejecución BORRA y reconstruye estas 8 pestañas. Si has editado
manualmente alguna celda, la perderás. Cuando los números estén firmes,
folder en rebuild_xlsx.py.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

WB_PATH = 'NEB-ALCARRIA-PRESUPUESTO.xlsx'

# Estilos
BOLD     = Font(bold=True)
BOLD_BIG = Font(bold=True, size=12)
BLUE     = Font(color='0000FF')
GREEN_T  = Font(color='008000')
ITALIC   = Font(italic=True, color='595959')
FILL_TITLE   = PatternFill('solid', start_color='8EA9DB')
FILL_HEADER  = PatternFill('solid', start_color='D9E1F2')
FILL_SUBHDR  = PatternFill('solid', start_color='FFE699')
FILL_LIGHTGRN= PatternFill('solid', start_color='C6EFCE')
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER   = Alignment(horizontal='center', vertical='center')

# ----------------------------------------------------------------------
# DATOS POR LL
# ----------------------------------------------------------------------

LLS = [
    {
        'id': 'LL0', 'idx': 0,
        'titulo': 'Concepto NEB + horticultura morisca alcarreña',
        'dias': 2, 'ediciones': 1, 'asistentes': 20,
        'periodo': 'P1', 'ventana': 'ene 2027',
        # Apoyo: (h_p_upm_GP1, h_p_upm_GP5, h_p_copr_GP1, h_p_copr_GP5)
        'apoyo': (16, 8, 16, 8),
        'apoyo_desc_upm_gp1': 'Técnico UPM. Dossier de bienvenida, materiales conceptuales NEB, gestión del ponente externo.',
        'apoyo_desc_upm_gp5': 'Oficial UPM. Montaje audiovisual, recepción de participantes.',
        'apoyo_desc_copr_gp1': 'Técnico territorial COPR. Coordinación con sabios del lugar y Ayuntamiento.',
        'apoyo_desc_copr_gp5': 'Oficial COPR. Montaje sala, manutención local, traslados internos.',
        # Docente: (h_doc, h_prep) -- pero los exponemos juntos como h/persona
        # Default = 8 h × días × (1 + 1,5) = 8 × 2 × 2,5 = 40
        'doc_upm_gp1': 40, 'doc_upm_gp5': 0, 'doc_copr_gp1': 40, 'doc_copr_gp5': 0,
        'doc_desc_upm_gp1': 'Tutor UPM. Marco NEB Compass, dinámica multiactor, conducción de los 2 días.',
        'doc_desc_copr_gp1': 'Tutor COPRODELI. Saber morisco hortícola, visita parcela, vínculo con sabios del lugar.',
        # Coordinación: (h_upm_gp1, h_upm_gp5, h_copr_gp1, h_copr_gp5)
        'coord': (8, 0, 8, 0),
        # Referencia para B.1 (fila en VIAJES)
        'viajes_row': 5,
    },
    {
        'id': 'LL1', 'idx': 1,
        'titulo': 'CUE+SIG · Itinerario digital del hortelano (cuaderno COPRODELI-IA + integración CUE/SIGPAC)',
        'dias': 10, 'ediciones': 1, 'asistentes': 27,
        'periodo': 'P1', 'ventana': 'feb–abr 2027',
        # 10 días con mucho trabajo previo de integración técnica (cuaderno COPRODELI-IA con CUE/SIEX/SIGPAC).
        # Ese trabajo NO es docencia, va a APOYO (sin cap RD-10).
        'apoyo': (100, 16, 60, 8),
        'apoyo_desc_upm_gp1': 'Ingeniero UPM. Integración del cuaderno COPRODELI-IA con CUE/SIEX/SIGPAC, conectores, plantillas QGIS/QField, despliegue en servidor.',
        'apoyo_desc_upm_gp5': 'Oficial UPM. Logística de las 2 semanas hábiles del LL.',
        'apoyo_desc_copr_gp1': 'Ingeniero COPRODELI. Adaptación del cuaderno COPRODELI-IA al contexto alcarreño, entrenamiento adicional de la IA con datos locales.',
        'apoyo_desc_copr_gp5': 'Oficial COPR. Manutención y logística in-situ.',
        # Docencia: 8 h × 10 días = 80 + prep 1,5× = 120 → 200 h/tutor (cap RD-10 al máximo en 1ª edición)
        'doc_upm_gp1': 200, 'doc_upm_gp5': 0, 'doc_copr_gp1': 200, 'doc_copr_gp5': 0,
        'doc_desc_upm_gp1': 'Tutor UPM. Imparte CUE/SIEX/SIGPAC, integración técnica, dronística para regadío.',
        'doc_desc_copr_gp1': 'Tutor COPRODELI. Imparte el módulo IA del cuaderno COPRODELI-IA, planificación de cultivos, apoyo decisional.',
        'coord': (40, 0, 40, 0),
        'viajes_row': 6,
    },
    {
        'id': 'LL2', 'idx': 2,
        'titulo': 'Drones agrarios · RGB (cartografía + regadío) + multiespectral (NDVI fitosanitario)',
        'dias': 4, 'ediciones': 1, 'asistentes': 25,
        'periodo': 'P2', 'ventana': 'sep 2027',
        # Apoyo previo: calibración dron, plan de vuelo, procesado WebODM
        'apoyo': (50, 10, 8, 4),
        'apoyo_desc_upm_gp1': 'Ingeniero UPM. Calibración multiespectral, plan de vuelo, procesado WebODM, plantillas NDVI/NDRE en QGIS.',
        'apoyo_desc_upm_gp5': 'Oficial UPM. Logística de campo, baterías, EPIs.',
        'apoyo_desc_copr_gp1': 'Técnico COPR. Coordinación territorial, permisos de vuelo locales.',
        'apoyo_desc_copr_gp5': 'Oficial COPR. Logística in-situ.',
        # 8 h × 4 días = 32 doc + 1,5× = 48 prep → 80 h/tutor
        'doc_upm_gp1': 80, 'doc_upm_gp5': 0, 'doc_copr_gp1': 80, 'doc_copr_gp5': 0,
        'doc_desc_upm_gp1': 'Tutor UPM. Vuelos RGB + multiespectral, procesado y análisis fitosanitario.',
        'doc_desc_copr_gp1': 'Tutor COPR. Conexión con realidad de las acequias y vegas alcarreñas, interpretación local.',
        'coord': (16, 0, 16, 0),
        'viajes_row': 7,
    },
    {
        'id': 'LL3', 'idx': 3,
        'titulo': 'Domo NEB · diseño digital + co-construcción (CAD → cálculo → estructura → cubierta → corcho)',
        'dias': 10, 'ediciones': 1, 'asistentes': 25,
        'periodo': 'P2', 'ventana': 'jul 2027',
        # Apoyo: precorte, marcado, andamios (oficial GP5 mucho peso)
        'apoyo': (30, 100, 30, 100),
        'apoyo_desc_upm_gp1': 'IP estructural UPM. Cálculo del polígono geodésico, supervisión técnica.',
        'apoyo_desc_upm_gp5': 'Oficial UPM. Precorte y marcado en taller.',
        'apoyo_desc_copr_gp1': 'Técnico COPR. Coordinación territorial con carpinteros locales y proveedores.',
        'apoyo_desc_copr_gp5': 'Oficial COPR. Carpintería local, precorte, montaje in-situ.',
        # 8 h × 10 días = 80 doc + 120 prep = 200 h/tutor
        'doc_upm_gp1': 200, 'doc_upm_gp5': 0, 'doc_copr_gp1': 200, 'doc_copr_gp5': 0,
        'doc_desc_upm_gp1': 'Tutor UPM. CAD, cálculo geodésico, supervisión estructural.',
        'doc_desc_copr_gp1': 'Maestro carpintero COPR. Técnicas constructivas con madera local, ensamblaje.',
        'coord': (40, 0, 40, 0),
        'viajes_row': 8,
    },
    {
        'id': 'LL4', 'idx': 4,
        'titulo': 'Living Lab digital · sensórica IoT + fotovoltaica + comunicaciones + volcado al SIG',
        'dias': 10, 'ediciones': 1, 'asistentes': 23,
        'periodo': 'P2', 'ventana': 'ago–sep 2027',
        # Apoyo: instalación y configuración intensiva (UPM lleva electrónica)
        'apoyo': (150, 80, 30, 40),
        'apoyo_desc_upm_gp1': 'Ingeniero UPM. Configuración red LoRa, gateway, integración con cuaderno COPRODELI-IA, dashboards.',
        'apoyo_desc_upm_gp5': 'Oficial UPM. Instalación física de sensores en parcela, cableado.',
        'apoyo_desc_copr_gp1': 'Técnico COPR. Coordinación con electricista local certificado BT.',
        'apoyo_desc_copr_gp5': 'Oficial COPR. Apoyo a instalación FV y montaje de antenas.',
        'doc_upm_gp1': 200, 'doc_upm_gp5': 0, 'doc_copr_gp1': 200, 'doc_copr_gp5': 0,
        'doc_desc_upm_gp1': 'Tutor UPM. Sensores LoRa, fotovoltaica, comunicaciones, integración SIG.',
        'doc_desc_copr_gp1': 'Tutor COPR. Conexión con realidad operativa de la microexplotación.',
        'coord': (40, 0, 40, 0),
        'viajes_row': 9,
    },
    {
        'id': 'LL5', 'idx': 5,
        'titulo': 'Ciclo productivo estacional · 4 sesiones (otoño inicio · otoño cosecha · primavera inicio · primavera cosecha)',
        'dias': 2, 'ediciones': 4, 'asistentes': 20,
        'periodo': 'P2', 'ventana': 'sep 2027 · dic 2027 · feb 2028 · abr 2028',
        # COPR lleva la huerta — peso en COPR
        'apoyo': (30, 40, 60, 80),
        'apoyo_desc_upm_gp1': 'Ingeniero UPM. Configuración IoT seguimiento, dashboards de ciclo, integración con IA del cuaderno.',
        'apoyo_desc_upm_gp5': 'Oficial UPM. Apoyo puntual a las 4 sesiones.',
        'apoyo_desc_copr_gp1': 'Agrónomo COPR. Planificación de los 2 ciclos productivos (otoño + primavera), siembra, abonado, riego.',
        'apoyo_desc_copr_gp5': 'Oficial agrícola COPR. Mantenimiento de la huerta entre sesiones.',
        # 4 ediciones × 2 días × 8 h = 64 h doc. Prep solo 1ª ed: 1,5 × 16 = 24 h. Total: 88 h/tutor.
        'doc_upm_gp1': 88, 'doc_upm_gp5': 0, 'doc_copr_gp1': 88, 'doc_copr_gp5': 0,
        'doc_desc_upm_gp1': 'Tutor UPM. Lectura de datos sensores, análisis IA cuaderno, NDVI dron.',
        'doc_desc_copr_gp1': 'Tutor COPR. Agronomía de la huerta morisca, cosecha, transferencia al hortelano.',
        'coord': (24, 0, 24, 0),
        'viajes_row': 10,
    },
    {
        'id': 'LL6', 'idx': 6,
        'titulo': 'Jornada final NEB-Alcarria · resultados, AKIS y replicabilidad',
        'dias': 1, 'ediciones': 1, 'asistentes': 40,
        'periodo': 'P2', 'ventana': 'abr 2028',
        'apoyo': (10, 5, 10, 5),
        'apoyo_desc_upm_gp1': 'Técnico UPM. Preparación dossier final, vídeo, materiales de difusión.',
        'apoyo_desc_upm_gp5': 'Oficial UPM. Montaje sala, recepción.',
        'apoyo_desc_copr_gp1': 'Técnico COPR. Convocatoria territorial, prensa local, autoridades.',
        'apoyo_desc_copr_gp5': 'Oficial COPR. Logística in-situ, manutención.',
        # 8 h × 1 día = 8 doc + 12 prep = 20 h/tutor
        'doc_upm_gp1': 20, 'doc_upm_gp5': 0, 'doc_copr_gp1': 20, 'doc_copr_gp5': 0,
        'doc_desc_upm_gp1': 'Tutor UPM. Presentación de resultados técnicos, métricas TRL alcanzadas.',
        'doc_desc_copr_gp1': 'Tutor COPR. Presentación de impacto territorial, replicabilidad.',
        'coord': (8, 0, 8, 0),
        'viajes_row': 11,
    },
]

# ----------------------------------------------------------------------
# UTILIDADES
# ----------------------------------------------------------------------

def coste_formula(num_col, grp_col, hrs_col, r):
    """Coste = Nº × h/persona × tarifa(GP1..GP5)"""
    return (f'={num_col}{r}*{hrs_col}{r}'
            f'*IF({grp_col}{r}="GP1",gp1_ss,'
            f'IF({grp_col}{r}="GP2",gp2_ss,'
            f'IF({grp_col}{r}="GP3",gp3_ss,'
            f'IF({grp_col}{r}="GP4",gp4_ss,'
            f'IF({grp_col}{r}="GP5",gp5_ss,0)))))')


def build_ll_sheet(wb, ll):
    """Construye una pestaña LLn con la estructura unificada."""
    sheet_name = ll['id']
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, index=4 + ll['idx'])

    widths = {'A': 34, 'B': 5, 'C': 8, 'D': 11, 'E': 11, 'F': 38,
              'G': 5, 'H': 8, 'I': 11, 'J': 11, 'K': 38}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Título
    ws['A1'] = f"{ll['id']} · {ll['titulo']}"
    ws['A1'].font = BOLD_BIG
    ws['A2'] = 'Hoja exploratoria. Convención: 1 día LL = 8 h imputables · h preparación ≤ 1,5 × h docencia (cap RD-10, solo docente).'
    ws['A2'].font = ITALIC

    # Sección 1 · Datos
    ws['A4'] = 'DATOS GENERALES'
    ws['A4'].font = BOLD; ws['A4'].fill = FILL_TITLE
    ws.merge_cells('A4:K4')

    datos = [
        ('Días',        ll['dias'],       'duración total del LL (input)'),
        ('Ediciones',   ll['ediciones'],  'nº de ediciones (input)'),
        ('Asistentes',  ll['asistentes'], 'total asistentes por edición'),
        ('Periodo',     ll['periodo'],    'P1 o P2'),
        ('Ventana',     ll['ventana'],    'mes(es) objetivo'),
    ]
    for i, (k, v, nota) in enumerate(datos):
        r = 5 + i
        ws.cell(row=r, column=1, value=k).font = BOLD
        c = ws.cell(row=r, column=2, value=v); c.font = BLUE
        ws.cell(row=r, column=4, value=nota).font = ITALIC

    # Sección 2 · Personal
    ws['A11'] = 'PERSONAL'
    ws['A11'].font = BOLD; ws['A11'].fill = FILL_TITLE
    ws.merge_cells('A11:K11')

    # Subcabecera UPM | COPR
    ws.cell(row=12, column=2, value='UPM').font = BOLD
    ws.cell(row=12, column=2).fill = FILL_HEADER
    ws.cell(row=12, column=2).alignment = CENTER
    ws.merge_cells('B12:F12')
    ws.cell(row=12, column=7, value='COPRODELI').font = BOLD
    ws.cell(row=12, column=7).fill = FILL_HEADER
    ws.cell(row=12, column=7).alignment = CENTER
    ws.merge_cells('G12:K12')

    headers = ['Concepto / Rol',
               'Nº', 'Grupo', 'h/persona', 'Coste', 'Descripción',
               'Nº', 'Grupo', 'h/persona', 'Coste', 'Descripción']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=13, column=i, value=h)
        c.font = BOLD; c.fill = FILL_HEADER; c.alignment = CENTER

    # --- 2.1 Personal de apoyo: 2 filas (GP1, GP5) ---
    ws.cell(row=14, column=1, value='Personal de apoyo').font = BOLD
    ws.cell(row=14, column=1).fill = FILL_SUBHDR
    ws.merge_cells('A14:K14')

    apoyo_upm_gp1, apoyo_upm_gp5, apoyo_copr_gp1, apoyo_copr_gp5 = ll['apoyo']

    # Fila GP1
    r = 15
    ws.cell(row=r, column=1, value='Apoyo · GP1 (ingeniero / técnico)').alignment = WRAP_TOP
    write_role_row(ws, r,
        upm_n=1 if apoyo_upm_gp1 > 0 else 0, upm_grp='GP1', upm_hp=apoyo_upm_gp1,
        upm_desc=ll['apoyo_desc_upm_gp1'],
        copr_n=1 if apoyo_copr_gp1 > 0 else 0, copr_grp='GP1', copr_hp=apoyo_copr_gp1,
        copr_desc=ll['apoyo_desc_copr_gp1'])
    # Fila GP5
    r = 16
    ws.cell(row=r, column=1, value='Apoyo · GP5 (oficial)').alignment = WRAP_TOP
    write_role_row(ws, r,
        upm_n=1 if apoyo_upm_gp5 > 0 else 0, upm_grp='GP5', upm_hp=apoyo_upm_gp5,
        upm_desc=ll['apoyo_desc_upm_gp5'],
        copr_n=1 if apoyo_copr_gp5 > 0 else 0, copr_grp='GP5', copr_hp=apoyo_copr_gp5,
        copr_desc=ll['apoyo_desc_copr_gp5'])

    # --- 2.2 Personal docente: 2 filas (GP1, GP5) ---
    ws.cell(row=18, column=1, value='Personal docente').font = BOLD
    ws.cell(row=18, column=1).fill = FILL_SUBHDR
    ws.merge_cells('A18:K18')

    r = 19
    ws.cell(row=r, column=1, value=f"Docente · GP1 (tutor presente todo el LL · {ll['dias']}d × 8h doc + 1,5× prep)").alignment = WRAP_TOP
    write_role_row(ws, r,
        upm_n=1 if ll['doc_upm_gp1'] > 0 else 0, upm_grp='GP1', upm_hp=ll['doc_upm_gp1'],
        upm_desc=ll['doc_desc_upm_gp1'],
        copr_n=1 if ll['doc_copr_gp1'] > 0 else 0, copr_grp='GP1', copr_hp=ll['doc_copr_gp1'],
        copr_desc=ll['doc_desc_copr_gp1'])
    r = 20
    ws.cell(row=r, column=1, value='Docente · GP5 (no aplica salvo casos especiales)').alignment = WRAP_TOP
    write_role_row(ws, r,
        upm_n=0, upm_grp='GP5', upm_hp=ll['doc_upm_gp5'], upm_desc='',
        copr_n=0, copr_grp='GP5', copr_hp=ll['doc_copr_gp5'], copr_desc='')

    # --- 2.3 Personal de coordinación: 2 filas (GP1, GP5) ---
    ws.cell(row=22, column=1, value='Personal de coordinación').font = BOLD
    ws.cell(row=22, column=1).fill = FILL_SUBHDR
    ws.merge_cells('A22:K22')

    coord_upm_gp1, coord_upm_gp5, coord_copr_gp1, coord_copr_gp5 = ll['coord']

    r = 23
    ws.cell(row=r, column=1, value='Coordinación · GP1 (imputación parcial al LL desde A.5)').alignment = WRAP_TOP
    write_role_row(ws, r,
        upm_n=1 if coord_upm_gp1 > 0 else 0, upm_grp='GP1', upm_hp=coord_upm_gp1,
        upm_desc='Coordinadora UPM · dedicación parcial al LL.',
        copr_n=1 if coord_copr_gp1 > 0 else 0, copr_grp='GP1', copr_hp=coord_copr_gp1,
        copr_desc='Coordinadora COPRODELI · dedicación parcial al LL.')
    r = 24
    ws.cell(row=r, column=1, value='Coordinación · GP5 (no aplica)').alignment = WRAP_TOP
    write_role_row(ws, r,
        upm_n=0, upm_grp='GP5', upm_hp=coord_upm_gp5, upm_desc='',
        copr_n=0, copr_grp='GP5', copr_hp=coord_copr_gp5, copr_desc='')

    # Sección 3 · RESUMEN
    ws['A26'] = f"RESUMEN — HORAS Y COSTES DE {ll['id']}"
    ws['A26'].font = BOLD; ws['A26'].fill = FILL_TITLE
    ws.merge_cells('A26:K26')

    # Headers
    ws.cell(row=27, column=1, value='Bloque').font = BOLD
    ws.cell(row=27, column=1).fill = FILL_HEADER
    for col_start, col_end, label in [(2,3,'Horas UPM'), (4,5,'Coste UPM'),
                                       (7,8,'Horas COPR'), (9,10,'Coste COPR')]:
        c = ws.cell(row=27, column=col_start, value=label)
        c.font = BOLD; c.fill = FILL_HEADER; c.alignment = CENTER
        ws.merge_cells(start_row=27, start_column=col_start, end_row=27, end_column=col_end)

    # Apoyo (filas 15-16), Docente (19-20), Coord (23-24)
    bloques = [
        ('Personal de apoyo',        [15, 16]),
        ('Personal docente',         [19, 20]),
        ('Personal de coordinación', [23, 24]),
    ]
    block_rows = []
    for i, (label, rows) in enumerate(bloques):
        r = 28 + i
        block_rows.append(r)
        ws.cell(row=r, column=1, value=label)
        # Horas UPM (suma Nº × h/persona en filas de UPM cols B,D)
        h_upm = "+".join(f'B{rr}*D{rr}' for rr in rows)
        ws.cell(row=r, column=2, value=f'={h_upm}')
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        # Coste UPM (suma columna E)
        c_upm = "+".join(f'E{rr}' for rr in rows)
        ws.cell(row=r, column=4, value=f'={c_upm}')
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        ws.cell(row=r, column=4).number_format = '#,##0.00 €'
        # Horas COPR (cols G,I)
        h_copr = "+".join(f'G{rr}*I{rr}' for rr in rows)
        ws.cell(row=r, column=7, value=f'={h_copr}')
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        # Coste COPR (col J)
        c_copr = "+".join(f'J{rr}' for rr in rows)
        ws.cell(row=r, column=9, value=f'={c_copr}')
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=10)
        ws.cell(row=r, column=9).number_format = '#,##0.00 €'

    # TOTAL PERSONAL LL
    tr = 31
    ws.cell(row=tr, column=1, value=f"TOTAL PERSONAL {ll['id']}").font = BOLD
    ws.cell(row=tr, column=2, value=f'=B{block_rows[0]}+B{block_rows[1]}+B{block_rows[2]}')
    ws.merge_cells(start_row=tr, start_column=2, end_row=tr, end_column=3)
    ws.cell(row=tr, column=4, value=f'=D{block_rows[0]}+D{block_rows[1]}+D{block_rows[2]}')
    ws.merge_cells(start_row=tr, start_column=4, end_row=tr, end_column=5)
    ws.cell(row=tr, column=4).number_format = '#,##0.00 €'
    ws.cell(row=tr, column=7, value=f'=G{block_rows[0]}+G{block_rows[1]}+G{block_rows[2]}')
    ws.merge_cells(start_row=tr, start_column=7, end_row=tr, end_column=8)
    ws.cell(row=tr, column=9, value=f'=I{block_rows[0]}+I{block_rows[1]}+I{block_rows[2]}')
    ws.merge_cells(start_row=tr, start_column=9, end_row=tr, end_column=10)
    ws.cell(row=tr, column=9).number_format = '#,##0.00 €'
    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K']:
        ws[f'{col_letter}{tr}'].fill = FILL_LIGHTGRN
        ws[f'{col_letter}{tr}'].font = BOLD

    # TOTAL COSTE UPM+COPR
    total_r = 32
    ws.cell(row=total_r, column=1, value='TOTAL COSTE PERSONAL (UPM + COPR)').font = BOLD
    ws.cell(row=total_r, column=4, value=f'=D{tr}+I{tr}')
    ws.merge_cells(start_row=total_r, start_column=4, end_row=total_r, end_column=5)
    ws.cell(row=total_r, column=4).font = BOLD
    ws.cell(row=total_r, column=4).number_format = '#,##0.00 €'
    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K']:
        ws[f'{col_letter}{total_r}'].fill = FILL_LIGHTGRN

    # B.1 referencia
    ws.cell(row=34, column=1,
            value=f"Para referencia · B.1 Viajes y dietas de {ll['id']} (calculado en VIAJES persona-día):"
           ).font = ITALIC
    ws.merge_cells(f'A34:I34')
    ws.cell(row=34, column=10, value=f"=VIAJES!P{ll['viajes_row']}").font = GREEN_T
    ws.cell(row=34, column=10).number_format = '#,##0.00 €'

    # TOTAL LL (personal + B.1)
    ws.cell(row=35, column=1, value=f"TOTAL {ll['id']} (personal + B.1)").font = BOLD
    ws.cell(row=35, column=10, value=f'=D{total_r}+J34').font = BOLD
    ws.cell(row=35, column=10).number_format = '#,##0.00 €'
    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K']:
        ws[f'{col_letter}35'].fill = FILL_LIGHTGRN

    # Nota externa
    ws.cell(row=37, column=1,
            value='Nota · Personal docente externo (A.1) y subcontratación (A.2) NO se incluyen aquí. Se añadirán al consolidar el modelo.'
           ).font = ITALIC
    ws.merge_cells('A37:K37')


def write_role_row(ws, r, upm_n, upm_grp, upm_hp, upm_desc, copr_n, copr_grp, copr_hp, copr_desc):
    """Escribe una fila de rol con UPM y COPR side-by-side."""
    c = ws.cell(row=r, column=2, value=upm_n); c.font = BLUE; c.alignment = CENTER
    c = ws.cell(row=r, column=3, value=upm_grp); c.alignment = CENTER
    c = ws.cell(row=r, column=4, value=upm_hp); c.font = BLUE; c.alignment = CENTER
    c = ws.cell(row=r, column=5, value=coste_formula('B','C','D',r)); c.number_format = '#,##0.00 €'
    ws.cell(row=r, column=6, value=upm_desc).alignment = WRAP_TOP

    c = ws.cell(row=r, column=7, value=copr_n); c.font = BLUE; c.alignment = CENTER
    c = ws.cell(row=r, column=8, value=copr_grp); c.alignment = CENTER
    c = ws.cell(row=r, column=9, value=copr_hp); c.font = BLUE; c.alignment = CENTER
    c = ws.cell(row=r, column=10, value=coste_formula('G','H','I',r)); c.number_format = '#,##0.00 €'
    ws.cell(row=r, column=11, value=copr_desc).alignment = WRAP_TOP


def build_resumen_sheet(wb):
    """Hoja RESUMEN — agrega costes de las 7 pestañas LL."""
    if 'RESUMEN_LLS' in wb.sheetnames:
        del wb['RESUMEN_LLS']
    ws = wb.create_sheet('RESUMEN_LLS', index=4 + len(LLS))

    widths = {'A': 7, 'B': 50, 'C': 6, 'D': 8, 'E': 9, 'F': 8,
              'G': 11, 'H': 11, 'I': 13, 'J': 13, 'K': 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws['A1'] = 'RESUMEN GENERAL · costes por Living Lab'
    ws['A1'].font = BOLD_BIG
    ws['A2'] = 'Agrega las cifras de las pestañas LL0..LL6. Los costes de PERSONAL son los que están en las pestañas LLn. El B.1 viene de la hoja VIAJES (modelo persona-día).'
    ws['A2'].font = ITALIC

    headers = ['Id', 'Título', 'Días', 'Edic.', 'Asist.', 'Periodo',
               'h UPM', 'h COPR', 'Coste personal', 'Coste B.1', 'TOTAL LL']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = BOLD; c.fill = FILL_HEADER; c.alignment = CENTER

    for i, ll in enumerate(LLS):
        r = 5 + i
        ws.cell(row=r, column=1, value=ll['id']).font = BOLD
        ws.cell(row=r, column=2, value=ll['titulo']).alignment = WRAP_TOP
        ws.cell(row=r, column=3, value=ll['dias']).alignment = CENTER
        ws.cell(row=r, column=4, value=ll['ediciones']).alignment = CENTER
        ws.cell(row=r, column=5, value=ll['asistentes']).alignment = CENTER
        ws.cell(row=r, column=6, value=ll['periodo']).alignment = CENTER
        # Referencias a cada LLn
        ws.cell(row=r, column=7, value=f"={ll['id']}!B31").font = GREEN_T  # h UPM (TOTAL PERSONAL row 31)
        ws.cell(row=r, column=8, value=f"={ll['id']}!G31").font = GREEN_T  # h COPR
        ws.cell(row=r, column=9, value=f"={ll['id']}!D32").font = GREEN_T  # coste personal total
        ws.cell(row=r, column=9).number_format = '#,##0.00 €'
        ws.cell(row=r, column=10, value=f"={ll['id']}!J34").font = GREEN_T  # B.1
        ws.cell(row=r, column=10).number_format = '#,##0.00 €'
        ws.cell(row=r, column=11, value=f"=I{r}+J{r}")
        ws.cell(row=r, column=11).number_format = '#,##0.00 €'
        ws.cell(row=r, column=11).font = BOLD

    # TOTAL fila
    tr = 5 + len(LLS)
    ws.cell(row=tr, column=2, value='TOTAL 7 Living Labs').font = BOLD
    for col in [7, 8, 9, 10, 11]:
        col_letter = get_column_letter(col)
        ws.cell(row=tr, column=col, value=f'=SUM({col_letter}5:{col_letter}{tr-1})')
        ws.cell(row=tr, column=col).number_format = '#,##0.00 €' if col >= 9 else 'General'
        ws.cell(row=tr, column=col).font = BOLD
    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K']:
        ws[f'{col_letter}{tr}'].fill = FILL_LIGHTGRN

    # Nota
    ws.cell(row=tr+2, column=1,
            value='Nota · Esta tabla refleja SOLO los costes desglosados por LL en las pestañas LLn (personal + B.1). El presupuesto formal del proyecto (con A.5 transversal, A.1/A.2 externos, B.2-B.7, C.1, C.2 e indirectos) sigue en la pestaña PRESUPUESTO.'
           ).font = ITALIC
    ws.merge_cells(f'A{tr+2}:K{tr+2}')


def main():
    wb = load_workbook(WB_PATH)
    print(f'Generando 7 pestañas LL + RESUMEN_LLS en {WB_PATH}...')
    for ll in LLS:
        build_ll_sheet(wb, ll)
        print(f'  ✓ {ll["id"]}')
    build_resumen_sheet(wb)
    print('  ✓ RESUMEN_LLS')
    wb.save(WB_PATH)
    print(f'\nGuardado. Pestañas:')
    print('  ' + ' · '.join(wb.sheetnames))


if __name__ == '__main__':
    main()
