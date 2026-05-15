"""
Reconstruye NEB-ALCARRIA-PRESUPUESTO.xlsx tras la migración al modelo de
Living Labs (LL-0 .. LL-6) con cuaderno COPRODELI-IA en TRL avanzado.

Estructura:
  - 7 Living Labs sustituyen las 25 actividades anteriores.
  - PERSONAL: A.1, A.2, A.3, A.4 (7 filas por LL) y A.5 (coordinación, 4 filas).
  - VIAJES: modelo persona-día (días con manutención completa, pernoctas,
    medias dietas) por LL — refleja la realidad de un LL multidía.
  - PRESUPUESTO: P1 (LL-0, LL-1) y P2 (LL-2..LL-6), con controles de topes.
  - MATERIALES e INVERSIONES: mismos importes; refs actualizadas a LL-x.

Uso:
  python3 build/rebuild_xlsx.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

WB_PATH = 'NEB-ALCARRIA-PRESUPUESTO.xlsx'

BLUE      = Font(color='0000FF')
GREEN_T   = Font(color='008000')
BOLD      = Font(bold=True)
BOLD_BIG  = Font(bold=True, size=12)
ITALIC    = Font(italic=True, color='595959')

FILL_YELLOW    = PatternFill('solid', start_color='FFFF00')
FILL_LIGHTGRN  = PatternFill('solid', start_color='C6EFCE')
FILL_HEADER    = PatternFill('solid', start_color='D9E1F2')
FILL_SECTION   = PatternFill('solid', start_color='8EA9DB')

WRAP = Alignment(wrap_text=True, vertical='top')
RIGHT = Alignment(horizontal='right', vertical='center')
CENTER = Alignment(horizontal='center', vertical='center')

# ----------------------------------------------------------------------
# DATOS MAESTROS — 7 Living Labs
# ----------------------------------------------------------------------

LLS = [
    {
        'id': 'LL-0',
        'familia': 'F0 · NEB + Morisca',
        'titulo': 'Concepto NEB + horticultura morisca alcarreña',
        'dias': 2, 'ediciones': 1,
        'periodo': 'P1', 'ventana': 'ene 2027',
        'asist_prof': 15, 'asist_estud': 5,
        'tut_upm': 2, 'tut_copr': 2,
        'pct_upm': 0.50,
        # VIAJES — persona-día
        'pd_manut': 48, 'pd_noches': 22, 'pd_media': 0,
        'vehiculos_x_viajes': 2,    # nº vehículos COPRODELI × nº viajes I+V grupo
        # A.4 horas
        'a4_gp1_upm': 4,  'a4_gp5_upm': 2,  'a4_gp1_copr': 4,  'a4_gp5_copr': 2,
        'a4_est_upm': 6,  'a4_est_copr': 0,
        # A.3 horas (docencia + preparación 1.5×)
        'a3_doc_upm': 6,  'a3_doc_copr': 8,  'a3_prep_upm': 9,  'a3_prep_copr': 12,
    },
    {
        'id': 'LL-1',
        'familia': 'F1 · CUE+SIG + COPRODELI-IA',
        'titulo': 'CUE+SIG · Itinerario digital del hortelano alcarreño (cuaderno COPRODELI-IA + integración CUE/SIGPAC)',
        'dias': 10, 'ediciones': 1,
        'periodo': 'P1', 'ventana': 'feb–abr 2027',
        'asist_prof': 22, 'asist_estud': 5,
        'tut_upm': 2, 'tut_copr': 2,
        'pct_upm': 0.65,
        # 27 viajeros · asist. media 5 d/persona = 145 + 2 locales × 10 d = 20 → 165 pd manut
        # 4 pernoctas/persona × 29 = 116 pd noches
        # 2 vehículos × 2 viajes I+V (una por semana) = 4 vehículo-viajes
        'pd_manut': 165, 'pd_noches': 116, 'pd_media': 0,
        'vehiculos_x_viajes': 4,
        # A.4 — grueso de la integración del cuaderno COPRODELI-IA con CUE/SIEX/SIGPAC
        'a4_gp1_upm': 250, 'a4_gp5_upm': 30, 'a4_gp1_copr': 100, 'a4_gp5_copr': 10,
        'a4_est_upm': 40, 'a4_est_copr': 0,
        'a3_doc_upm': 25, 'a3_doc_copr': 20, 'a3_prep_upm': 60, 'a3_prep_copr': 40,
    },
    {
        'id': 'LL-2',
        'familia': 'F2 · Drones agrarios',
        'titulo': 'Drones agrarios · RGB (cartografía + regadío) + multiespectral (NDVI fitosanitario)',
        'dias': 4, 'ediciones': 1,
        'periodo': 'P2', 'ventana': 'sep 2027',
        'asist_prof': 18, 'asist_estud': 7,
        'tut_upm': 2, 'tut_copr': 1,
        'pct_upm': 0.75,
        # 27 viajeros × 4 d = 108 + 1 local × 4 = 4 → 112 pd manut
        # 27 × 3 = 81 noches
        'pd_manut': 112, 'pd_noches': 81, 'pd_media': 0,
        'vehiculos_x_viajes': 2,
        'a4_gp1_upm': 60, 'a4_gp5_upm': 10, 'a4_gp1_copr': 8, 'a4_gp5_copr': 4,
        'a4_est_upm': 20, 'a4_est_copr': 0,
        'a3_doc_upm': 12, 'a3_doc_copr': 4, 'a3_prep_upm': 30, 'a3_prep_copr': 6,
    },
    {
        'id': 'LL-3',
        'familia': 'F3 · Domo + Diseño',
        'titulo': 'Domo NEB · diseño digital + co-construcción (CAD → cálculo → estructura → cubierta → corcho)',
        'dias': 10, 'ediciones': 1,
        'periodo': 'P2', 'ventana': 'jul 2027',
        'asist_prof': 15, 'asist_estud': 10,
        'tut_upm': 2, 'tut_copr': 3,
        'pct_upm': 0.35,
        # 27 viajeros × 7 d asist. = 189 + 3 locales × 10 = 30 → 219 pd manut
        # 27 × 6 = 162 noches
        'pd_manut': 219, 'pd_noches': 162, 'pd_media': 0,
        'vehiculos_x_viajes': 4,
        'a4_gp1_upm': 30, 'a4_gp5_upm': 100, 'a4_gp1_copr': 30, 'a4_gp5_copr': 100,
        'a4_est_upm': 60, 'a4_est_copr': 0,
        'a3_doc_upm': 20, 'a3_doc_copr': 20, 'a3_prep_upm': 40, 'a3_prep_copr': 40,
    },
    {
        'id': 'LL-4',
        'familia': 'F4 · Sensórica + FV + Com.',
        'titulo': 'Living Lab digital · sensórica IoT + fotovoltaica + comunicaciones + volcado al SIG',
        'dias': 10, 'ediciones': 1,
        'periodo': 'P2', 'ventana': 'ago–sep 2027',
        'asist_prof': 15, 'asist_estud': 8,
        'tut_upm': 2, 'tut_copr': 2,
        'pct_upm': 0.65,
        # 25 × 7 = 175 + 2 × 10 = 20 → 195 pd manut
        # 25 × 6 = 150 noches
        'pd_manut': 195, 'pd_noches': 150, 'pd_media': 0,
        'vehiculos_x_viajes': 4,
        'a4_gp1_upm': 150, 'a4_gp5_upm': 80, 'a4_gp1_copr': 30, 'a4_gp5_copr': 40,
        'a4_est_upm': 40, 'a4_est_copr': 0,
        'a3_doc_upm': 30, 'a3_doc_copr': 8, 'a3_prep_upm': 60, 'a3_prep_copr': 12,
    },
    {
        'id': 'LL-5',
        'familia': 'F5 · Ciclo productivo',
        'titulo': 'Ciclo productivo estacional · 4 sesiones (otoño inicio · otoño cosecha · primavera inicio · primavera cosecha)',
        'dias': 2, 'ediciones': 4,
        'periodo': 'P2', 'ventana': 'sep 2027 · dic 2027 · feb 2028 · abr 2028',
        'asist_prof': 15, 'asist_estud': 5,
        'tut_upm': 1, 'tut_copr': 2,
        'pct_upm': 0.30,
        # Por sesión: 21 viajeros × 2 d = 42 + 2 locales × 2 = 4 → 46 pd manut
        # × 4 sesiones = 184 pd manut
        # Pernoctas: 21 × 1 × 4 = 84
        # Vehículos: 2 × 1 I+V × 4 sesiones = 8
        'pd_manut': 184, 'pd_noches': 84, 'pd_media': 0,
        'vehiculos_x_viajes': 8,
        'a4_gp1_upm': 30, 'a4_gp5_upm': 40, 'a4_gp1_copr': 60, 'a4_gp5_copr': 80,
        'a4_est_upm': 40, 'a4_est_copr': 0,
        'a3_doc_upm': 8, 'a3_doc_copr': 20, 'a3_prep_upm': 12, 'a3_prep_copr': 30,
    },
    {
        'id': 'LL-6',
        'familia': 'F6 · Cierre',
        'titulo': 'Jornada final NEB-Alcarria · resultados, AKIS y replicabilidad',
        'dias': 1, 'ediciones': 1,
        'periodo': 'P2', 'ventana': 'abr 2028',
        'asist_prof': 30, 'asist_estud': 10,
        'tut_upm': 2, 'tut_copr': 2,
        'pct_upm': 0.50,
        # 1 día, sin pernocta — todos almuerzan: 42 viajeros + 2 locales = 44 medias dietas
        'pd_manut': 0, 'pd_noches': 0, 'pd_media': 44,
        'vehiculos_x_viajes': 3,
        'a4_gp1_upm': 10, 'a4_gp5_upm': 5, 'a4_gp1_copr': 10, 'a4_gp5_copr': 5,
        'a4_est_upm': 6, 'a4_est_copr': 0,
        'a3_doc_upm': 4, 'a3_doc_copr': 4, 'a3_prep_upm': 6, 'a3_prep_copr': 6,
    },
]

A1_ROWS = [
    # Sin ponentes externos previstos en ALMUNIA 4.0.
    # Se mantiene una fila placeholder con 0 h para preservar la estructura del bloque A.1
    # (las fórmulas SUMIF de PRESUPUESTO siguen funcionando y A.1 total = 0 €).
    ('—', 'No procede · sin ponentes externos en el proyecto', 0, 0),
    ('—', 'No procede · sin ponentes externos en el proyecto', 0, 0),
]

A2_ROWS = [
    ('LL-3', 'Carpintero externo · precorte madera del domo',     15, 35),
    ('LL-3', 'Albañil externo · preinstalación de 16 anclajes',    9, 21),
    ('LL-3', 'Carpintero externo · prefijación de cubierta',       9, 21),
    ('LL-4', 'Instalador eléctrico · pre-cableado FV',             8,  8),
]

A5_ROWS = [
    ('Coordinadora UPM — P1',       'UPM',       'P1', 32, 'meses_p1'),
    ('Coordinadora UPM — P2',       'UPM',       'P2', 25, 'meses_p2'),
    ('Coordinadora COPRODELI — P1', 'COPRODELI', 'P1', 32, 'meses_p1'),
    ('Coordinadora COPRODELI — P2', 'COPRODELI', 'P2', 25, 'meses_p2'),
]

# ----------------------------------------------------------------------
# UTILIDADES
# ----------------------------------------------------------------------

def clear_sheet(ws):
    """Borra contenido, estilos y celdas combinadas de una hoja."""
    merged = list(ws.merged_cells.ranges)
    for rng in merged:
        ws.unmerge_cells(str(rng))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
            cell.number_format = 'General'

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write_section_title(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = BOLD_BIG
    ws.cell(row=row, column=1).fill = FILL_SECTION

def write_headers(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = BOLD
        c.fill = FILL_HEADER
        c.alignment = WRAP

def style_total(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = FILL_LIGHTGRN

def style_input(cell):
    cell.font = BLUE

def style_xref(cell):
    cell.font = GREEN_T

def style_hyp(cell):
    cell.fill = FILL_YELLOW

# ----------------------------------------------------------------------
# ACTIVIDADES
# ----------------------------------------------------------------------

def build_actividades(ws):
    clear_sheet(ws)
    set_widths(ws, {'A': 8, 'B': 28, 'C': 70, 'D': 7, 'E': 10,
                    'F': 14, 'G': 14, 'H': 12, 'I': 10, 'J': 26})

    ws.cell(row=1, column=1, value='ACTIVIDADES — listado maestro · 7 Living Labs (LL-0..LL-6)').font = BOLD_BIG
    ws.cell(row=2, column=1, value='Sustituye el modelo anterior de 25 jornadas sueltas. Cada LL es una actividad demostrativa presencial unitaria (Art. 5.6 RD 251/2024).').font = ITALIC

    headers = ['Id', 'Familia', 'Actividad', 'Días', 'Ediciones',
               'Asist. prof./ed.', 'Asist. estud./ed.', 'Asist. total', 'Periodo', 'Ventana']
    write_headers(ws, 4, headers)

    start_row = 5
    for i, ll in enumerate(LLS):
        r = start_row + i
        ws.cell(row=r, column=1, value=ll['id'])
        ws.cell(row=r, column=2, value=ll['familia'])
        ws.cell(row=r, column=3, value=ll['titulo']).alignment = WRAP
        c = ws.cell(row=r, column=4, value=ll['dias']);      style_input(c)
        c = ws.cell(row=r, column=5, value=ll['ediciones']); style_input(c)
        c = ws.cell(row=r, column=6, value=ll['asist_prof']);  style_input(c)
        c = ws.cell(row=r, column=7, value=ll['asist_estud']); style_input(c)
        ws.cell(row=r, column=8, value=f'=(F{r}+G{r})*E{r}')
        ws.cell(row=r, column=9, value=ll['periodo'])
        ws.cell(row=r, column=10, value=ll['ventana'])

    # TOTAL
    total_r = start_row + len(LLS)
    ws.cell(row=total_r, column=3, value='TOTAL')
    ws.cell(row=total_r, column=4, value=f'=SUMPRODUCT(D{start_row}:D{total_r-1},E{start_row}:E{total_r-1})')  # días-actividad
    ws.cell(row=total_r, column=5, value=f'=SUM(E{start_row}:E{total_r-1})')
    ws.cell(row=total_r, column=8, value=f'=SUM(H{start_row}:H{total_r-1})')
    style_total(ws, total_r, 10)

# ----------------------------------------------------------------------
# PERSONAL
# ----------------------------------------------------------------------

# Layout (rows):
#   R4  A.4 section title
#   R5  headers
#   R6..R12  LL-0..LL-6 (7 filas)
#   R13 TOTAL A.4
#
#   R16 A.3 title
#   R17 headers
#   R18..R24 LL-0..LL-6
#   R25 TOTAL A.3
#
#   R28 A.1 title
#   R29 headers
#   R30..R31 A.1 (2 filas)
#   R32 TOTAL A.1
#
#   R35 A.2 title
#   R36 headers
#   R37..R40 A.2 (4 filas)
#   R41 TOTAL A.2
#
#   R44 A.5 title
#   R45 headers
#   R46..R49 A.5 (4 filas)
#   R50 TOTAL A.5

# Estos índices se usan también para construir PRESUPUESTO
A4_FIRST, A4_LAST, A4_TOTAL = 6, 12, 13
A3_FIRST, A3_LAST, A3_TOTAL = 18, 24, 25
A1_FIRST, A1_LAST, A1_TOTAL = 30, 31, 32
A2_FIRST, A2_LAST, A2_TOTAL = 37, 40, 41
A5_FIRST, A5_LAST, A5_TOTAL = 46, 49, 50

def build_personal(ws):
    clear_sheet(ws)
    set_widths(ws, {'A': 8, 'B': 38, 'C': 11, 'D': 11, 'E': 11, 'F': 11,
                    'G': 11, 'H': 11, 'I': 12, 'J': 12, 'K': 12})

    ws.cell(row=1, column=1, value='PERSONAL — A.1 · A.2 · A.3 · A.4 · A.5').font = BOLD_BIG
    ws.cell(row=2, column=1, value='Inputs en azul. Tarifas €/h se toman de PARAMETROS. Estructura por Living Lab (LL-0..LL-6).').font = ITALIC

    # ----------------- A.4 PERSONAL DE APOYO PROPIO -----------------
    write_section_title(ws, 4, 'A.4 · PERSONAL DE APOYO PROPIO · acondicionamiento Living Lab + integración del cuaderno COPRODELI-IA con CUE/SIEX/SIGPAC')
    write_headers(ws, 5, ['Id', 'Familia', 'h GP1 UPM', 'h GP1 COPR', 'h GP5 UPM', 'h GP5 COPR',
                          'h Estud. UPM', 'h Estud. COPR', 'Coste UPM', 'Coste COPR', 'Coste total'])
    for i, ll in enumerate(LLS):
        r = A4_FIRST + i
        ws.cell(row=r, column=1, value=ll['id'])
        ws.cell(row=r, column=2, value=f"=VLOOKUP(A{r},ACTIVIDADES!$A$5:$B$50,2,FALSE())").font = GREEN_T
        for col, key in [(3,'a4_gp1_upm'),(4,'a4_gp1_copr'),(5,'a4_gp5_upm'),
                         (6,'a4_gp5_copr'),(7,'a4_est_upm'),(8,'a4_est_copr')]:
            c = ws.cell(row=r, column=col, value=ll[key]); style_input(c)
        ws.cell(row=r, column=9,  value=f'=C{r}*gp1_ss+E{r}*gp5_ss')   # UPM
        ws.cell(row=r, column=10, value=f'=D{r}*gp1_ss+F{r}*gp5_ss')   # COPR
        ws.cell(row=r, column=11, value=f'=I{r}+J{r}')
        ws.cell(row=r, column=9).number_format = '#,##0.00'
        ws.cell(row=r, column=10).number_format = '#,##0.00'
        ws.cell(row=r, column=11).number_format = '#,##0.00'

    # TOTAL A.4
    ws.cell(row=A4_TOTAL, column=2, value='TOTAL A.4')
    for col in [3,4,5,6,7,8,9,10,11]:
        ws.cell(row=A4_TOTAL, column=col,
                value=f'=SUM({get_column_letter(col)}{A4_FIRST}:{get_column_letter(col)}{A4_LAST})')
    for col in [9,10,11]:
        ws.cell(row=A4_TOTAL, column=col).number_format = '#,##0.00'
    style_total(ws, A4_TOTAL, 11)

    # ----------------- A.3 PERSONAL DOCENTE PROPIO -----------------
    write_section_title(ws, 16, 'A.3 · PERSONAL DOCENTE PROPIO · GP1 ponentes UPM / COPRODELI · ratio preparación 1,5× sobre h impartida')
    write_headers(ws, 17, ['Id', 'Familia', 'h doc UPM', 'h doc COPR', 'h prep UPM', 'h prep COPR',
                           'Total h UPM', 'Total h COPR', 'Coste UPM', 'Coste COPR', 'Coste total'])
    for i, ll in enumerate(LLS):
        r = A3_FIRST + i
        ws.cell(row=r, column=1, value=ll['id'])
        ws.cell(row=r, column=2, value=f"=VLOOKUP(A{r},ACTIVIDADES!$A$5:$B$50,2,FALSE())").font = GREEN_T
        for col, key in [(3,'a3_doc_upm'),(4,'a3_doc_copr'),(5,'a3_prep_upm'),(6,'a3_prep_copr')]:
            c = ws.cell(row=r, column=col, value=ll[key]); style_input(c)
        ws.cell(row=r, column=7, value=f'=C{r}+E{r}')
        ws.cell(row=r, column=8, value=f'=D{r}+F{r}')
        ws.cell(row=r, column=9,  value=f'=G{r}*gp1_ss')
        ws.cell(row=r, column=10, value=f'=H{r}*gp1_ss')
        ws.cell(row=r, column=11, value=f'=I{r}+J{r}')
        for col in [9,10,11]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'
    ws.cell(row=A3_TOTAL, column=2, value='TOTAL A.3')
    for col in [3,4,5,6,7,8,9,10,11]:
        ws.cell(row=A3_TOTAL, column=col,
                value=f'=SUM({get_column_letter(col)}{A3_FIRST}:{get_column_letter(col)}{A3_LAST})')
    for col in [9,10,11]:
        ws.cell(row=A3_TOTAL, column=col).number_format = '#,##0.00'
    style_total(ws, A3_TOTAL, 11)

    # ----------------- A.1 PERSONAL DOCENTE EXTERNO -----------------
    write_section_title(ws, 28, 'A.1 · PERSONAL DOCENTE EXTERNO · ponentes invitados · 90 €/h (IVA excl.)')
    write_headers(ws, 29, ['Id', 'Concepto', 'h ext. UPM', 'h ext. COPR', 'Total h',
                           'Coste UPM', 'Coste COPR', 'Coste total'])
    for i, (ll_id, concepto, h_upm, h_copr) in enumerate(A1_ROWS):
        r = A1_FIRST + i
        ws.cell(row=r, column=1, value=ll_id)
        ws.cell(row=r, column=2, value=concepto)
        c = ws.cell(row=r, column=3, value=h_upm);  style_input(c)
        c = ws.cell(row=r, column=4, value=h_copr); style_input(c)
        ws.cell(row=r, column=5, value=f'=C{r}+D{r}')
        ws.cell(row=r, column=6, value=f'=C{r}*ext_h')
        ws.cell(row=r, column=7, value=f'=D{r}*ext_h')
        ws.cell(row=r, column=8, value=f'=F{r}+G{r}')
        for col in [6,7,8]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'
    ws.cell(row=A1_TOTAL, column=2, value='TOTAL A.1')
    for col in [3,4,5,6,7,8]:
        ws.cell(row=A1_TOTAL, column=col,
                value=f'=SUM({get_column_letter(col)}{A1_FIRST}:{get_column_letter(col)}{A1_LAST})')
    for col in [6,7,8]:
        ws.cell(row=A1_TOTAL, column=col).number_format = '#,##0.00'
    style_total(ws, A1_TOTAL, 8)

    # ----------------- A.2 PERSONAL DE APOYO EXTERNO -----------------
    write_section_title(ws, 35, 'A.2 · PERSONAL DE APOYO EXTERNO · subcontratados (carpintería, albañilería, electricidad) · 90 €/h')
    write_headers(ws, 36, ['Id', 'Concepto', 'h ext. UPM', 'h ext. COPR', 'Total h',
                           'Coste UPM', 'Coste COPR', 'Coste total'])
    for i, (ll_id, concepto, h_upm, h_copr) in enumerate(A2_ROWS):
        r = A2_FIRST + i
        ws.cell(row=r, column=1, value=ll_id)
        ws.cell(row=r, column=2, value=concepto)
        c = ws.cell(row=r, column=3, value=h_upm);  style_input(c)
        c = ws.cell(row=r, column=4, value=h_copr); style_input(c)
        ws.cell(row=r, column=5, value=f'=C{r}+D{r}')
        ws.cell(row=r, column=6, value=f'=C{r}*ext_h')
        ws.cell(row=r, column=7, value=f'=D{r}*ext_h')
        ws.cell(row=r, column=8, value=f'=F{r}+G{r}')
        for col in [6,7,8]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'
    ws.cell(row=A2_TOTAL, column=2, value='TOTAL A.2')
    for col in [3,4,5,6,7,8]:
        ws.cell(row=A2_TOTAL, column=col,
                value=f'=SUM({get_column_letter(col)}{A2_FIRST}:{get_column_letter(col)}{A2_LAST})')
    for col in [6,7,8]:
        ws.cell(row=A2_TOTAL, column=col).number_format = '#,##0.00'
    style_total(ws, A2_TOTAL, 8)

    # ----------------- A.5 COORDINACIÓN PROPIA -----------------
    write_section_title(ws, 44, 'A.5 · PERSONAL DE COORDINACIÓN PROPIO · UPM + COPRODELI · GP1 · desglose por fase · tope 20 % directos')
    write_headers(ws, 45, ['Persona', 'Entidad', 'Periodo', 'h/mes', 'Meses', 'Total h', '€/h', 'Coste'])
    for i, (persona, ent, per, h_mes, meses_name) in enumerate(A5_ROWS):
        r = A5_FIRST + i
        ws.cell(row=r, column=1, value=persona)
        ws.cell(row=r, column=2, value=ent)
        ws.cell(row=r, column=3, value=per)
        c = ws.cell(row=r, column=4, value=h_mes); style_input(c)
        ws.cell(row=r, column=5, value=f'={meses_name}').font = GREEN_T
        ws.cell(row=r, column=6, value=f'=D{r}*E{r}')
        ws.cell(row=r, column=7, value='=gp1_ss').font = GREEN_T
        ws.cell(row=r, column=8, value=f'=F{r}*G{r}')
        ws.cell(row=r, column=8).number_format = '#,##0.00'
    ws.cell(row=A5_TOTAL, column=2, value='TOTAL A.5')
    ws.cell(row=A5_TOTAL, column=4, value=f'=SUM(D{A5_FIRST}:D{A5_LAST})')
    ws.cell(row=A5_TOTAL, column=6, value=f'=SUM(F{A5_FIRST}:F{A5_LAST})')
    ws.cell(row=A5_TOTAL, column=8, value=f'=SUM(H{A5_FIRST}:H{A5_LAST})')
    ws.cell(row=A5_TOTAL, column=8).number_format = '#,##0.00'
    style_total(ws, A5_TOTAL, 8)

# ----------------------------------------------------------------------
# VIAJES — modelo persona-día
# ----------------------------------------------------------------------

VIAJES_FIRST = 5
VIAJES_LAST  = 11
VIAJES_TOTAL = 12
VIAJES_P1    = 13
VIAJES_P2    = 14

def build_viajes(ws):
    clear_sheet(ws)
    set_widths(ws, {
        'A': 7,  'B': 30, 'C': 6, 'D': 9, 'E': 9, 'F': 8, 'G': 8,
        'H': 9, 'I': 12, 'J': 11, 'K': 14, 'L': 14,
        'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 8, 'R': 11, 'S': 11, 'T': 8
    })

    ws.cell(row=1, column=1, value='VIAJES — B.1 · Transporte + alojamiento + manutención · modelo persona-día').font = BOLD_BIG
    ws.cell(row=2, column=1, value='Una fila por Living Lab. Precios COPRODELI (PARAMETROS bloque E) por debajo de los topes oficiales. Manutención/aloj. cuentan persona-día; transporte cuenta vehículo-viaje I+V (vehículos COPRODELI 20 plazas).').font = ITALIC

    headers = ['Id', 'Familia', 'Días LL', 'Asist. prof', 'Asist. estud', 'Tutor UPM', 'Tutor COPR',
               'Viajeros (D+E+F)', 'Manut. completa (pd)', 'Pernoctas (pn)', 'Media dieta (pd)',
               'Vehíc.×viajes I+V', 'Manutención €', 'Alojamiento €', 'Transporte €', 'Total €',
               'UPM %', 'UPM €', 'COPR €', 'Periodo']
    write_headers(ws, 4, headers)

    for i, ll in enumerate(LLS):
        r = VIAJES_FIRST + i
        ws.cell(row=r, column=1, value=ll['id'])
        ws.cell(row=r, column=2, value=ll['familia'])
        c = ws.cell(row=r, column=3, value=ll['dias']);        style_input(c)
        c = ws.cell(row=r, column=4, value=ll['asist_prof']);  style_input(c)
        c = ws.cell(row=r, column=5, value=ll['asist_estud']); style_input(c)
        c = ws.cell(row=r, column=6, value=ll['tut_upm']);     style_input(c)
        c = ws.cell(row=r, column=7, value=ll['tut_copr']);    style_input(c)
        ws.cell(row=r, column=8, value=f'=D{r}+E{r}+F{r}')
        c = ws.cell(row=r, column=9,  value=ll['pd_manut']);  style_input(c); style_hyp(c)
        c = ws.cell(row=r, column=10, value=ll['pd_noches']); style_input(c); style_hyp(c)
        c = ws.cell(row=r, column=11, value=ll['pd_media']);  style_input(c); style_hyp(c)
        c = ws.cell(row=r, column=12, value=ll['vehiculos_x_viajes']); style_input(c); style_hyp(c)
        # Manutención €: persona-días × manut completa precio COPR + media dieta × media COPR
        ws.cell(row=r, column=13, value=f'=I{r}*manut_dia_copr+K{r}*manut_med_copr')
        ws.cell(row=r, column=14, value=f'=J{r}*alojamiento_copr')
        # Transporte € = vehículos × 2 trayectos × km × €/km
        ws.cell(row=r, column=15, value=f'=L{r}*2*km_madrid_pastrana*km_coche')
        ws.cell(row=r, column=16, value=f'=M{r}+N{r}+O{r}')
        c = ws.cell(row=r, column=17, value=ll['pct_upm']); style_input(c)
        ws.cell(row=r, column=17).number_format = '0%'
        ws.cell(row=r, column=18, value=f'=P{r}*Q{r}')
        ws.cell(row=r, column=19, value=f'=P{r}*(1-Q{r})')
        ws.cell(row=r, column=20, value=ll['periodo'])
        for col in [13,14,15,16,18,19]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    # TOTAL
    ws.cell(row=VIAJES_TOTAL, column=2, value='TOTAL B.1')
    for col in [4,5,6,7,9,10,11,12,13,14,15,16,18,19]:
        ws.cell(row=VIAJES_TOTAL, column=col,
                value=f'=SUM({get_column_letter(col)}{VIAJES_FIRST}:{get_column_letter(col)}{VIAJES_LAST})')
    for col in [13,14,15,16,18,19]:
        ws.cell(row=VIAJES_TOTAL, column=col).number_format = '#,##0.00'
    style_total(ws, VIAJES_TOTAL, 20)
    # Subtotales P1 y P2
    ws.cell(row=VIAJES_P1, column=2, value='Subtotal P1')
    ws.cell(row=VIAJES_P1, column=16, value=f'=SUMIF(T{VIAJES_FIRST}:T{VIAJES_LAST},"P1",P{VIAJES_FIRST}:P{VIAJES_LAST})')
    ws.cell(row=VIAJES_P1, column=16).number_format = '#,##0.00'
    style_total(ws, VIAJES_P1, 20)

    ws.cell(row=VIAJES_P2, column=2, value='Subtotal P2')
    ws.cell(row=VIAJES_P2, column=16, value=f'=SUMIF(T{VIAJES_FIRST}:T{VIAJES_LAST},"P2",P{VIAJES_FIRST}:P{VIAJES_LAST})')
    ws.cell(row=VIAJES_P2, column=16).number_format = '#,##0.00'
    style_total(ws, VIAJES_P2, 20)

# ----------------------------------------------------------------------
# MATERIALES — actualización de descripciones (sin cambio estructural)
# ----------------------------------------------------------------------

MATERIALES_ROWS = [
    # (Partida, Concepto, Cantidad, € unit, Tope, UPM %, Periodo)
    ('B.2', 'Entornos virtuales · plataforma online complementaria (LMS de recursos y materiales)',
        1, 1000, 'tope_ent', 0.50, 'P1'),
    ('B.3', 'Material fungible P1 · 30 participantes únicos × 35 €/u (cuadernos, USB, dossier)',
        30, 35, '—', 0.40, 'P1'),
    ('B.3', 'Material fungible P2 · 50 participantes únicos × 35 €/u',
        50, 35, '—', 0.40, 'P2'),
    ('B.4', 'Seguros P1 · RC actividades demostrativas + accidentes',
        1, 400, '—', 0.50, 'P1'),
    ('B.4', 'Seguros P2 · RC actividades demostrativas + accidentes',
        1, 600, '—', 0.50, 'P2'),
    ('B.5', 'Auditoría · auditor externo (1 % subvención · tope 5.000 €)',
        1, 2800, 'tope_aud', 0.50, 'P2'),
    ('B.6', 'Alquileres P1 · salas formación, herramientas puntuales',
        1, 1500, '—', 0.50, 'P1'),
    ('B.6', 'Alquileres P2 · andamios LL-3 (domo), herramientas LL-4 (sensórica)',
        1, 3500, '—', 0.50, 'P2'),
    ('B.7', 'Comunicación P1 · web, RRSS, podcast inicial',
        1, 500, 'tope_com', 0.50, 'P1'),
    ('B.7', 'Comunicación P2 · vídeo, roll-ups, dossier final',
        1, 1000, 'tope_com', 0.50, 'P2'),
]

def build_materiales(ws):
    clear_sheet(ws)
    set_widths(ws, {'A': 8, 'B': 60, 'C': 9, 'D': 9, 'E': 11, 'F': 10, 'G': 8, 'H': 8, 'I': 10, 'J': 9})

    ws.cell(row=1, column=1, value='MATERIALES — B.2 · B.3 · B.4 · B.5 · B.6 · B.7').font = BOLD_BIG
    ws.cell(row=2, column=1, value='Cada concepto con su Periodo (input azul). Topes calculados con PARAMETROS.').font = ITALIC

    write_headers(ws, 4, ['Partida', 'Concepto', 'Cantidad', '€ unit.', 'Coste', 'Tope', '% tope',
                          'UPM %', 'UPM €', 'Periodo'])
    for i, (partida, concepto, cant, eu, tope, upm_pct, per) in enumerate(MATERIALES_ROWS):
        r = 5 + i
        ws.cell(row=r, column=1, value=partida)
        ws.cell(row=r, column=2, value=concepto).alignment = WRAP
        c = ws.cell(row=r, column=3, value=cant); style_input(c)
        c = ws.cell(row=r, column=4, value=eu);   style_input(c)
        ws.cell(row=r, column=5, value=f'=C{r}*D{r}')
        if tope == '—':
            ws.cell(row=r, column=6, value='—')
            ws.cell(row=r, column=7, value='—')
        else:
            ws.cell(row=r, column=6, value=f'={tope}').font = GREEN_T
            ws.cell(row=r, column=7, value=f'=E{r}/F{r}')
            ws.cell(row=r, column=7).number_format = '0%'
        c = ws.cell(row=r, column=8, value=upm_pct); style_input(c)
        ws.cell(row=r, column=8).number_format = '0%'
        ws.cell(row=r, column=9, value=f'=E{r}*H{r}')
        ws.cell(row=r, column=10, value=per)
        for col in [5,9]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    total_r = 5 + len(MATERIALES_ROWS)
    ws.cell(row=total_r, column=2, value='TOTAL B (sin B.1)')
    ws.cell(row=total_r, column=5, value=f'=SUM(E5:E{total_r-1})')
    ws.cell(row=total_r, column=9, value=f'=SUM(I5:I{total_r-1})')
    for col in [5,9]:
        ws.cell(row=total_r, column=col).number_format = '#,##0.00'
    style_total(ws, total_r, 10)

# ----------------------------------------------------------------------
# INVERSIONES — mismos importes, refs actualizadas a LL-x
# ----------------------------------------------------------------------

C1_ROWS = [
    # (Concepto, Descripción, Coste, Vida útil m, UPM %, Notas)
    ('Dron 4K', 'Cuerpo del dron 4K (proveedor distinto al de óptica multiespectral)',
        3500, 60, 1.0, 'LL-2 · activo UPM (drones GeoSo2). < 5.000 €'),
    ('Cámara multiespectral', 'Cámara multiespectral compatible con dron — sensor NDVI/NDRE',
        4500, 60, 1.0, 'LL-2 · activo UPM. Proveedor especializado distinto del dron. < 5.000 €'),
    ('Sensores IoT', 'Lote de 15 sensores LoRa humedad de suelo / meteo',
        3700, 60, 1.0, 'LL-4 (despliegue) + LL-5 (operación) · quedan en parcela · activo UPM. < 5.000 €'),
    ('Software de cálculo de estructuras', '3 licencias anuales (IP estructural + 2 colaboradores UPM)',
        3000, 17, 1.0, 'LL-3 · diseño domo. 3 licencias anuales prorrateadas al periodo'),
    ('Servidor / gateway IoT', 'Gateway LoRa + servidor backend (cuaderno COPRODELI-IA en producción + telemetría IoT)',
        2000, 60, 1.0, 'LL-1 (backend cuaderno) + LL-4 (gateway IoT) · activo UPM'),
]

C2_ROWS = [
    # (Concepto, Descripción, Coste, UPM %, Vinculación)
    ('Cimentación domo', 'Zapatas, hormigón armado, replanteo del polígono geodésico',
        2500, 0.30, 'LL-3'),
    ('Listones del domo', 'Madera precortada para los listones del domo (aserradero local)',
        3500, 0.30, 'LL-3'),
    ('Tornillería y herramientas del domo', 'Tornillería de precisión, herrajes metálicos (ferretería industrial)',
        1500, 0.30, 'LL-3'),
    ('Cubierta del domo', 'Cubierta exterior del domo (proveedor especializado de cubiertas técnicas)',
        4500, 0.30, 'LL-3'),
    ('Aislamiento térmico del domo', 'Aislamiento térmico con corcho local (bioeconomía circular)',
        3000, 0.30, 'LL-3'),
    ('Aislamiento hidrófugo del domo', 'Membrana hidrófuga de impermeabilización exterior',
        2500, 0.30, 'LL-3'),
    ('Ventanas del domo', 'Ventanas y elementos de ventilación pasiva',
        2500, 0.30, 'LL-3'),
    ('Paneles fotovoltaicos', 'Paneles FV bifaciales N-type 5 kW (importador especializado)',
        4500, 0.60, 'LL-4'),
    ('Baterías, cableado e instalación FV', 'Baterías + cableado + montaje BT (electricista local certificado)',
        3500, 0.60, 'LL-4'),
    ('Conectividad satelital', 'Starlink + soporte + cuotas 17 meses',
        2500, 0.50, 'LL-4'),
    ('Semillas y sustratos', 'Plantel y abonos · 2 ciclos productivos (otoño + primavera)',
        2000, 0.20, 'LL-5'),
    ('Sensórica y comunicaciones de riego', 'Sensores humedad de suelo + módulo LoRa para riego automatizado',
        1500, 0.70, 'LL-4 + LL-5'),
    ('Material de riego', 'Electroválvulas, tuberías, goteros, filtros',
        2000, 0.40, 'LL-4 + LL-5'),
]

def build_inversiones(ws):
    clear_sheet(ws)
    set_widths(ws, {'A': 28, 'B': 55, 'C': 11, 'D': 9, 'E': 11, 'F': 11, 'G': 11, 'H': 7, 'I': 11, 'J': 11, 'K': 45})

    ws.cell(row=1, column=1, value='INVERSIONES — C.1 (amortización bienes inventariables) · C.2 (otras inversiones)').font = BOLD_BIG

    # C.1
    write_section_title(ws, 3, 'C.1 · BIENES INVENTARIABLES — amortización en periodo subvencionable')
    write_headers(ws, 4, ['Concepto', 'Descripción', 'Coste compra', 'Vida útil (m)',
                          'Cuota P1 (7 m)', 'Cuota P2 (10 m)', 'Cuota total',
                          'UPM %', 'UPM €', 'COPR €', 'Notas'])
    c1_start = 5
    for i, (concepto, descr, coste, vida, upm_pct, notas) in enumerate(C1_ROWS):
        r = c1_start + i
        ws.cell(row=r, column=1, value=concepto)
        ws.cell(row=r, column=2, value=descr).alignment = WRAP
        c = ws.cell(row=r, column=3, value=coste); style_input(c)
        c = ws.cell(row=r, column=4, value=vida);  style_input(c)
        # Si vida útil >= meses_proy, cuota total = coste; sino prorrateo
        ws.cell(row=r, column=5, value=f'=C{r}*MIN(D{r},meses_p1)/D{r}')
        ws.cell(row=r, column=6, value=f'=C{r}*MIN(D{r},meses_p2)/D{r}')
        # Para soft estr (D=17 m), prorrateo da Coste*7/17 + Coste*10/17 = Coste; correcto.
        # Para hw (D=60), 7/60 + 10/60 = 17/60 → cuota total = Coste×17/60. Correcto.
        ws.cell(row=r, column=7, value=f'=E{r}+F{r}')
        c = ws.cell(row=r, column=8, value=upm_pct); style_input(c)
        ws.cell(row=r, column=8).number_format = '0%'
        ws.cell(row=r, column=9, value=f'=G{r}*H{r}')
        ws.cell(row=r, column=10, value=f'=G{r}*(1-H{r})')
        ws.cell(row=r, column=11, value=notas).alignment = WRAP
        for col in [3,5,6,7,9,10]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    c1_total = c1_start + len(C1_ROWS)
    ws.cell(row=c1_total, column=2, value='TOTAL C.1')
    for col in [5,6,7,9,10]:
        ws.cell(row=c1_total, column=col,
                value=f'=SUM({get_column_letter(col)}{c1_start}:{get_column_letter(col)}{c1_total-1})')
        ws.cell(row=c1_total, column=col).number_format = '#,##0.00'
    style_total(ws, c1_total, 11)
    global C1_TOTAL_ROW
    C1_TOTAL_ROW = c1_total

    # C.2
    c2_section = c1_total + 3
    write_section_title(ws, c2_section, 'C.2 · OTRAS INVERSIONES — cada ítem < 5.000 € (no requiere 3 ofertas en Fase I)')
    c2_header = c2_section + 1
    write_headers(ws, c2_header, ['Concepto', 'Descripción', 'Cantidad', '€ unit.', 'Coste total',
                                   'Periodo', 'UPM %', 'UPM €', 'COPR €', 'Vinculación'])
    c2_start = c2_header + 1
    for i, (concepto, descr, coste, upm_pct, vinc) in enumerate(C2_ROWS):
        r = c2_start + i
        ws.cell(row=r, column=1, value=concepto)
        ws.cell(row=r, column=2, value=descr).alignment = WRAP
        c = ws.cell(row=r, column=3, value=1); style_input(c)
        c = ws.cell(row=r, column=4, value=coste); style_input(c)
        ws.cell(row=r, column=5, value=f'=C{r}*D{r}')
        ws.cell(row=r, column=6, value='P2')
        c = ws.cell(row=r, column=7, value=upm_pct); style_input(c)
        ws.cell(row=r, column=7).number_format = '0%'
        ws.cell(row=r, column=8, value=f'=E{r}*G{r}')
        ws.cell(row=r, column=9, value=f'=E{r}*(1-G{r})')
        ws.cell(row=r, column=10, value=vinc)
        for col in [4,5,8,9]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    c2_total = c2_start + len(C2_ROWS)
    ws.cell(row=c2_total, column=2, value='TOTAL C.2')
    for col in [5,8,9]:
        ws.cell(row=c2_total, column=col,
                value=f'=SUM({get_column_letter(col)}{c2_start}:{get_column_letter(col)}{c2_total-1})')
        ws.cell(row=c2_total, column=col).number_format = '#,##0.00'
    style_total(ws, c2_total, 10)
    global C2_TOTAL_ROW, C2_FIRST, C2_LAST
    C2_TOTAL_ROW = c2_total
    C2_FIRST = c2_start
    C2_LAST = c2_total - 1

# ----------------------------------------------------------------------
# PRESUPUESTO — desglose por fase
# ----------------------------------------------------------------------

def build_presupuesto(ws):
    clear_sheet(ws)
    set_widths(ws, {'A': 8, 'B': 60, 'C': 14, 'D': 8, 'E': 14, 'F': 11, 'G': 14})

    ws.cell(row=1, column=1, value='PRESUPUESTO — desglose por fase y Living Lab').font = BOLD_BIG
    ws.cell(row=2, column=1, value='Cada Living Lab se imputa a P1 o P2 según el periodo de su primera edición. Costes transversales (A.5, B, C, indirectos) se asignan explícitamente a cada fase.').font = ITALIC

    # ----- FASE 1 (P1) -----
    write_section_title(ws, 4, 'FASE 1 (P1) — Living Labs y costes transversales')
    headers_p = ['Id', 'Actividad', 'A.1', 'A.2', 'A.3', 'A.4', 'Subtotal personal LL']
    write_headers(ws, 5, headers_p)

    p1_lls = [ll['id'] for ll in LLS if ll['periodo'] == 'P1']
    p1_first = 6
    for i, ll_id in enumerate(p1_lls):
        r = p1_first + i
        ws.cell(row=r, column=1, value=ll_id)
        ws.cell(row=r, column=2, value=f"=VLOOKUP(A{r},ACTIVIDADES!$A$5:$C$50,3,FALSE())").font = GREEN_T
        ws.cell(row=r, column=3, value=f"=SUMIF(PERSONAL!$A${A1_FIRST}:$A${A1_LAST},A{r},PERSONAL!$H${A1_FIRST}:$H${A1_LAST})").font = GREEN_T
        ws.cell(row=r, column=4, value=f"=SUMIF(PERSONAL!$A${A2_FIRST}:$A${A2_LAST},A{r},PERSONAL!$H${A2_FIRST}:$H${A2_LAST})").font = GREEN_T
        ws.cell(row=r, column=5, value=f"=SUMIF(PERSONAL!$A${A3_FIRST}:$A${A3_LAST},A{r},PERSONAL!$K${A3_FIRST}:$K${A3_LAST})").font = GREEN_T
        ws.cell(row=r, column=6, value=f"=SUMIF(PERSONAL!$A${A4_FIRST}:$A${A4_LAST},A{r},PERSONAL!$K${A4_FIRST}:$K${A4_LAST})").font = GREEN_T
        ws.cell(row=r, column=7, value=f"=C{r}+D{r}+E{r}+F{r}")
        for col in [3,4,5,6,7]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    p1_subtotal = p1_first + len(p1_lls)
    ws.cell(row=p1_subtotal, column=2, value='Subtotal personal P1')
    for col in [3,4,5,6,7]:
        ws.cell(row=p1_subtotal, column=col,
                value=f'=SUM({get_column_letter(col)}{p1_first}:{get_column_letter(col)}{p1_subtotal-1})')
        ws.cell(row=p1_subtotal, column=col).number_format = '#,##0.00'
    style_total(ws, p1_subtotal, 7)

    p1_trans_title = p1_subtotal + 2
    write_section_title(ws, p1_trans_title, 'COSTES TRANSVERSALES P1')
    write_headers(ws, p1_trans_title + 1, ['', 'Concepto', 'Total'])

    trans_p1_concepts = [
        ('A.5 Coordinación', f'=PERSONAL!$H${A5_FIRST}+PERSONAL!$H${A5_FIRST+2}'),  # UPM-P1 + COPR-P1
        ('B.1 Viajes y dietas (precios COPRODELI)', f'=VIAJES!$P${VIAJES_P1}'),
        ('B.2 Entornos virtuales', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.2",MATERIALES!$J$5:$J$14,"P1")'),
        ('B.3 Material fungible', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.3",MATERIALES!$J$5:$J$14,"P1")'),
        ('B.4 Seguros', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.4",MATERIALES!$J$5:$J$14,"P1")'),
        ('B.5 Auditoría', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.5",MATERIALES!$J$5:$J$14,"P1")'),
        ('B.6 Alquileres', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.6",MATERIALES!$J$5:$J$14,"P1")'),
        ('B.7 Comunicación', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.7",MATERIALES!$J$5:$J$14,"P1")'),
        ('C.1 Bienes inventariables (cuota P1)', f'=INVERSIONES!$E${C1_TOTAL_ROW}'),
        ('C.2 Otras inversiones (P1)', f'=SUMIFS(INVERSIONES!$E${C2_FIRST}:$E${C2_LAST},INVERSIONES!$F${C2_FIRST}:$F${C2_LAST},"P1")'),
    ]
    p1_trans_first = p1_trans_title + 2
    for i, (label, formula) in enumerate(trans_p1_concepts):
        r = p1_trans_first + i
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=formula).font = GREEN_T
        ws.cell(row=r, column=3).number_format = '#,##0.00'
    # Indirectos (15 % sobre personal directo de fase = subtotal personal P1 + A.5 P1)
    p1_indir = p1_trans_first + len(trans_p1_concepts)
    ws.cell(row=p1_indir, column=2, value='Costes indirectos (15 % personal directo de la fase)')
    ws.cell(row=p1_indir, column=3, value=f'=($G${p1_subtotal}+C{p1_trans_first})*tasa_ind')
    ws.cell(row=p1_indir, column=3).number_format = '#,##0.00'
    p1_subtot_trans = p1_indir + 1
    ws.cell(row=p1_subtot_trans, column=2, value='Subtotal transversales P1')
    ws.cell(row=p1_subtot_trans, column=3, value=f'=SUM(C{p1_trans_first}:C{p1_indir})')
    ws.cell(row=p1_subtot_trans, column=3).number_format = '#,##0.00'
    style_total(ws, p1_subtot_trans, 7)
    p1_total = p1_subtot_trans + 1
    ws.cell(row=p1_total, column=2, value='TOTAL P1')
    ws.cell(row=p1_total, column=3, value=f'=$G${p1_subtotal}+C{p1_subtot_trans}')
    ws.cell(row=p1_total, column=3).number_format = '#,##0.00'
    style_total(ws, p1_total, 7)

    # ----- FASE 2 (P2) -----
    p2_title_row = p1_total + 3
    write_section_title(ws, p2_title_row, 'FASE 2 (P2) — Living Labs y costes transversales')
    write_headers(ws, p2_title_row + 1, headers_p)

    p2_lls = [ll['id'] for ll in LLS if ll['periodo'] == 'P2']
    p2_first = p2_title_row + 2
    for i, ll_id in enumerate(p2_lls):
        r = p2_first + i
        ws.cell(row=r, column=1, value=ll_id)
        ws.cell(row=r, column=2, value=f"=VLOOKUP(A{r},ACTIVIDADES!$A$5:$C$50,3,FALSE())").font = GREEN_T
        ws.cell(row=r, column=3, value=f"=SUMIF(PERSONAL!$A${A1_FIRST}:$A${A1_LAST},A{r},PERSONAL!$H${A1_FIRST}:$H${A1_LAST})").font = GREEN_T
        ws.cell(row=r, column=4, value=f"=SUMIF(PERSONAL!$A${A2_FIRST}:$A${A2_LAST},A{r},PERSONAL!$H${A2_FIRST}:$H${A2_LAST})").font = GREEN_T
        ws.cell(row=r, column=5, value=f"=SUMIF(PERSONAL!$A${A3_FIRST}:$A${A3_LAST},A{r},PERSONAL!$K${A3_FIRST}:$K${A3_LAST})").font = GREEN_T
        ws.cell(row=r, column=6, value=f"=SUMIF(PERSONAL!$A${A4_FIRST}:$A${A4_LAST},A{r},PERSONAL!$K${A4_FIRST}:$K${A4_LAST})").font = GREEN_T
        ws.cell(row=r, column=7, value=f"=C{r}+D{r}+E{r}+F{r}")
        for col in [3,4,5,6,7]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    p2_subtotal = p2_first + len(p2_lls)
    ws.cell(row=p2_subtotal, column=2, value='Subtotal personal P2')
    for col in [3,4,5,6,7]:
        ws.cell(row=p2_subtotal, column=col,
                value=f'=SUM({get_column_letter(col)}{p2_first}:{get_column_letter(col)}{p2_subtotal-1})')
        ws.cell(row=p2_subtotal, column=col).number_format = '#,##0.00'
    style_total(ws, p2_subtotal, 7)

    p2_trans_title = p2_subtotal + 2
    write_section_title(ws, p2_trans_title, 'COSTES TRANSVERSALES P2')
    write_headers(ws, p2_trans_title + 1, ['', 'Concepto', 'Total'])
    trans_p2_concepts = [
        ('A.5 Coordinación', f'=PERSONAL!$H${A5_FIRST+1}+PERSONAL!$H${A5_FIRST+3}'),  # UPM-P2 + COPR-P2
        ('B.1 Viajes y dietas (precios COPRODELI)', f'=VIAJES!$P${VIAJES_P2}'),
        ('B.2 Entornos virtuales', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.2",MATERIALES!$J$5:$J$14,"P2")'),
        ('B.3 Material fungible', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.3",MATERIALES!$J$5:$J$14,"P2")'),
        ('B.4 Seguros', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.4",MATERIALES!$J$5:$J$14,"P2")'),
        ('B.5 Auditoría', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.5",MATERIALES!$J$5:$J$14,"P2")'),
        ('B.6 Alquileres', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.6",MATERIALES!$J$5:$J$14,"P2")'),
        ('B.7 Comunicación', '=SUMIFS(MATERIALES!$E$5:$E$14,MATERIALES!$A$5:$A$14,"B.7",MATERIALES!$J$5:$J$14,"P2")'),
        ('C.1 Bienes inventariables (cuota P2)', f'=INVERSIONES!$F${C1_TOTAL_ROW}'),
        ('C.2 Otras inversiones (P2)', f'=SUMIFS(INVERSIONES!$E${C2_FIRST}:$E${C2_LAST},INVERSIONES!$F${C2_FIRST}:$F${C2_LAST},"P2")'),
    ]
    p2_trans_first = p2_trans_title + 2
    for i, (label, formula) in enumerate(trans_p2_concepts):
        r = p2_trans_first + i
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=formula).font = GREEN_T
        ws.cell(row=r, column=3).number_format = '#,##0.00'
    p2_indir = p2_trans_first + len(trans_p2_concepts)
    ws.cell(row=p2_indir, column=2, value='Costes indirectos (15 % personal directo de la fase)')
    ws.cell(row=p2_indir, column=3, value=f'=($G${p2_subtotal}+C{p2_trans_first})*tasa_ind')
    ws.cell(row=p2_indir, column=3).number_format = '#,##0.00'
    p2_subtot_trans = p2_indir + 1
    ws.cell(row=p2_subtot_trans, column=2, value='Subtotal transversales P2')
    ws.cell(row=p2_subtot_trans, column=3, value=f'=SUM(C{p2_trans_first}:C{p2_indir})')
    ws.cell(row=p2_subtot_trans, column=3).number_format = '#,##0.00'
    style_total(ws, p2_subtot_trans, 7)
    p2_total = p2_subtot_trans + 1
    ws.cell(row=p2_total, column=2, value='TOTAL P2')
    ws.cell(row=p2_total, column=3, value=f'=$G${p2_subtotal}+C{p2_subtot_trans}')
    ws.cell(row=p2_total, column=3).number_format = '#,##0.00'
    style_total(ws, p2_total, 7)

    # ----- TOTALES GENERALES -----
    grand = p2_total + 2
    ws.cell(row=grand, column=2, value='TOTAL GENERAL (P1 + P2)')
    ws.cell(row=grand, column=3, value=f'=C{p1_total}+C{p2_total}')
    ws.cell(row=grand, column=3).number_format = '#,##0.00'
    ws.cell(row=grand, column=3).font = BOLD
    style_total(ws, grand, 7)

    directos_row = grand + 2
    ws.cell(row=directos_row, column=2, value='Costes directos totales (sin indirectos)')
    # Suma de todo el personal directo (A.1+A.2+A.3+A.4+A.5) + B.1 + B.2..B.7 + C.1 + C.2
    ws.cell(row=directos_row, column=3, value=(
        f'=PERSONAL!$H${A1_TOTAL}+PERSONAL!$H${A2_TOTAL}'
        f'+PERSONAL!$K${A3_TOTAL}+PERSONAL!$K${A4_TOTAL}+PERSONAL!$H${A5_TOTAL}'
        f'+VIAJES!$P${VIAJES_TOTAL}'
        f'+SUMIF(MATERIALES!$A$5:$A$14,"B.*",MATERIALES!$E$5:$E$14)'
        f'+INVERSIONES!$G${C1_TOTAL_ROW}+INVERSIONES!$E${C2_TOTAL_ROW}'
    ))
    # The SUMIF wildcard "B.*" doesn't match in Excel by default — use plain SUM of MATERIALES total instead
    ws.cell(row=directos_row, column=3, value=(
        f'=PERSONAL!$H${A1_TOTAL}+PERSONAL!$H${A2_TOTAL}'
        f'+PERSONAL!$K${A3_TOTAL}+PERSONAL!$K${A4_TOTAL}+PERSONAL!$H${A5_TOTAL}'
        f'+VIAJES!$P${VIAJES_TOTAL}'
        f'+SUM(MATERIALES!$E$5:$E$14)'
        f'+INVERSIONES!$G${C1_TOTAL_ROW}+INVERSIONES!$E${C2_TOTAL_ROW}'
    ))
    ws.cell(row=directos_row, column=3).number_format = '#,##0.00'

    # ----- CONTROL DE TOPES -----
    topes_title = directos_row + 3
    write_section_title(ws, topes_title, 'CONTROL DE TOPES')
    write_headers(ws, topes_title + 1, ['', 'Concepto', 'Valor', '', 'Tope', '%', 'Estado'])
    topes = [
        ('A.5 / directos ≤ 20 %',           f'=PERSONAL!$H${A5_TOTAL}',              f'=$C${directos_row}*tope_coord'),
        ('B.1 / directos ≤ 40 %',           f'=VIAJES!$P${VIAJES_TOTAL}',            f'=$C${directos_row}*tope_viajes'),
        ('(C.1+C.2) / directos ≤ 50 %',     f'=INVERSIONES!$G${C1_TOTAL_ROW}+INVERSIONES!$E${C2_TOTAL_ROW}', f'=$C${directos_row}*tope_inv'),
        ('B.2 ≤ 5.000 €',                   '=SUMIF(MATERIALES!$A$5:$A$14,"B.2",MATERIALES!$E$5:$E$14)', '=tope_ent'),
        ('B.4 ≤ 1.000 €',                   '=SUMIF(MATERIALES!$A$5:$A$14,"B.4",MATERIALES!$E$5:$E$14)', '=tope_seg'),
        ('B.7 ≤ 1.500 €',                   '=SUMIF(MATERIALES!$A$5:$A$14,"B.7",MATERIALES!$E$5:$E$14)', '=tope_com'),
        ('Total ≤ 500.000 €',               f'=$C${grand}',                          '=tope_subv'),
        ('Total ≥ 70.000 €',                f'=$C${grand}',                          '=min_subv'),
    ]
    for i, (label, val_formula, tope_formula) in enumerate(topes):
        r = topes_title + 2 + i
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=val_formula).font = GREEN_T
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=5, value=tope_formula).font = GREEN_T
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=f'=C{r}/E{r}')
        ws.cell(row=r, column=6).number_format = '0.0%'
        # Estado — para los dos últimos invertimos la comparación
        if 'Total ≥' in label:
            ws.cell(row=r, column=7, value=f'=IF(C{r}>=E{r},"OK","BAJO MIN")')
        else:
            ws.cell(row=r, column=7, value=f'=IF(C{r}<=E{r},"OK","SUPERADO")')

# ----------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------

def main():
    wb = load_workbook(WB_PATH)

    print(f'Reconstruyendo hojas en {WB_PATH}...')
    build_actividades(wb['ACTIVIDADES'])
    print('  ✓ ACTIVIDADES (7 LLs)')

    build_personal(wb['PERSONAL'])
    print('  ✓ PERSONAL (A.1..A.5 con 7 LLs)')

    build_viajes(wb['VIAJES'])
    print('  ✓ VIAJES (modelo persona-día)')

    build_materiales(wb['MATERIALES'])
    print('  ✓ MATERIALES')

    build_inversiones(wb['INVERSIONES'])
    print('  ✓ INVERSIONES (refs actualizadas a LL-x)')

    build_presupuesto(wb['PRESUPUESTO'])
    print('  ✓ PRESUPUESTO (P1: LL-0+LL-1 · P2: LL-2..LL-6)')

    wb.save(WB_PATH)
    print(f'\nGuardado: {WB_PATH}')
    print('Ejecuta scripts/recalc.py NEB-ALCARRIA-PRESUPUESTO.xlsx para verificar.')


if __name__ == '__main__':
    main()
