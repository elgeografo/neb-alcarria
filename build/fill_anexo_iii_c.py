"""Rellena el Anexo III_C (cronograma) con datos del proyecto NEB-ALCARRIA.

Fuente de verdad: hoja ACTIVIDADES de NEB-ALCARRIA-PRESUPUESTO.xlsx (7 Living Labs).
Periodos: P1 nov 2026 → may 2027 (7 m) · P2 jul 2027 → abr 2028 (10 m).

El template tiene 3 filas para actividades en cada periodo (11-13 y 16-18).
P1 usa 2 LL (LL-0, LL-1) + 1 fila de coordinación.
P2 necesita 5 LL (LL-2..LL-6), así que insertamos 2 filas tras 18.
"""
from copy import copy
from openpyxl import load_workbook

SRC = "docu a presentar/03_18_mayo/V03/Anexo_III_C_CRONOGRAMA_NEB-ALCARRIA.xlsx"

wb = load_workbook(SRC)
ws = wb["CRONOGRAMA"]

MONTH_COLS = ["D", "E", "F", "G", "H", "I", "J", "K", "L", "N", "O", "P"]

# --- A. Insertar 2 filas para LL-5 y LL-6 antes de que openpyxl trate de mover
#        las celdas combinadas (que no shifta bien). Para evitar problemas
#        primero desmergeamos los rangos afectados, hacemos la inserción y
#        después restablecemos los merges en sus nuevas posiciones. ---
FOOTER_RANGES_OLD = ["A19:P19", "A20:P20", "A21:P21"]
for r in FOOTER_RANGES_OLD:
    if r in ws.merged_cells:
        ws.unmerge_cells(r)

# Guardar valores y estilos de las filas 19-22 antes de mover.
def cell_snapshot(row):
    snap = []
    for col_idx in range(1, ws.max_column + 1):
        c = ws.cell(row=row, column=col_idx)
        snap.append({
            "value": c.value,
            "font": copy(c.font),
            "fill": copy(c.fill),
            "border": copy(c.border),
            "alignment": copy(c.alignment),
            "number_format": c.number_format,
        })
    return snap

snap_19 = cell_snapshot(19)
snap_20 = cell_snapshot(20)
snap_21 = cell_snapshot(21)
snap_22 = cell_snapshot(22)

height_19 = ws.row_dimensions[19].height
height_20 = ws.row_dimensions[20].height
height_21 = ws.row_dimensions[21].height
height_22 = ws.row_dimensions[22].height

# Limpiar filas 19-22
for r in (19, 20, 21, 22):
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=r, column=col_idx).value = None

# Clonar estilo de fila 18 (activity row) en filas 19 y 20.
src_row = 18
for new_row in (19, 20):
    ws.row_dimensions[new_row].height = ws.row_dimensions[src_row].height
    for col_idx in range(1, ws.max_column + 1):
        s = ws.cell(row=src_row, column=col_idx)
        d = ws.cell(row=new_row, column=col_idx)
        d.font = copy(s.font)
        d.fill = copy(s.fill)
        d.border = copy(s.border)
        d.alignment = copy(s.alignment)
        d.number_format = s.number_format

# Replicar merge L:M en filas 19, 20 (cabecera del mes "Sept" / "Mar" se solapa con L:M)
ws.merge_cells("L19:M19")
ws.merge_cells("L20:M20")

# Restaurar contenido de filas footer en filas 21-23 (desplazadas 2)
def apply_snapshot(target_row, snap, height):
    if height:
        ws.row_dimensions[target_row].height = height
    for col_idx, info in enumerate(snap, start=1):
        c = ws.cell(row=target_row, column=col_idx)
        c.value = info["value"]
        c.font = info["font"]
        c.fill = info["fill"]
        c.border = info["border"]
        c.alignment = info["alignment"]
        c.number_format = info["number_format"]

apply_snapshot(21, snap_19, height_19)
apply_snapshot(22, snap_20, height_20)
apply_snapshot(23, snap_21, height_21)
apply_snapshot(24, snap_22, height_22)

# Restaurar merges de footer en sus nuevas posiciones
ws.merge_cells("A21:P21")
ws.merge_cells("A22:P22")
ws.merge_cells("A23:P23")

# --- B. Cabecera ---
ws["B5"] = "ALMUNIA 4.0 (agrupación en formación) · Fundación COPRODELI + Universidad Politécnica de Madrid"
ws["K5"] = "No aplica (agrupación)"
ws["B7"] = "NEB-ALCARRIA — Living Lab demostrativo de horticultura digital de microexplotación en la Alcarria"
ws["I7"] = "Programa I.B"

# --- C. Etiquetas de meses ---
P1_MONTHS = [
    "Nov\n2026", "Dic\n2026", "Ene\n2027", "Feb\n2027", "Mar\n2027",
    "Abr\n2027", "May\n2027", "—", "—", "—", "—", "—",
]
P2_MONTHS = [
    "Jul\n2027", "Ago\n2027", "Sep\n2027", "Oct\n2027", "Nov\n2027",
    "Dic\n2027", "Ene\n2028", "Feb\n2028", "Mar\n2028", "Abr\n2028", "—", "—",
]

for col, label in zip(MONTH_COLS, P1_MONTHS):
    ws[f"{col}10"] = label
for col, label in zip(MONTH_COLS, P2_MONTHS):
    ws[f"{col}15"] = label

# --- D. Limpiar las casillas de meses en filas de actividad ---
ACT_ROWS_P1 = [11, 12, 13]
ACT_ROWS_P2 = [16, 17, 18, 19, 20]
for r in ACT_ROWS_P1 + ACT_ROWS_P2:
    for col in MONTH_COLS:
        ws[f"{col}{r}"] = "☐"

# --- E. Fusionar B:C de cada fila de actividad para que el nombre quepa ---
for r in ACT_ROWS_P1 + ACT_ROWS_P2:
    rng = f"B{r}:C{r}"
    if rng not in ws.merged_cells:
        ws.merge_cells(rng)

# --- F. Definiciones de actividades ---
def mark(row: int, idx: int, value: str):
    ws[f"{MONTH_COLS[idx]}{row}"] = value

# P1 actividades
ws["B11"] = "LL-0 · Concepto NEB + horticultura morisca alcarreña"
mark(11, 2, "E1")  # Ene 2027

ws["B12"] = "LL-1 · CUE+SIG · Itinerario digital del hortelano alcarreño (cuaderno COPRODELI-IA + integración CUE/SIGPAC)"
mark(12, 3, "E1")  # Feb 2027
mark(12, 4, "E1")  # Mar 2027
mark(12, 5, "E1")  # Abr 2027

ws["B13"] = "Coordinación y acondicionamiento previo del Living Lab (personal de apoyo) — continuo"
for i in range(7):
    mark(13, i, "●")

# Limpiar casillas inexistentes (índices 7-11) en P1
for r in (11, 12, 13):
    for i in range(7, 12):
        ws[f"{MONTH_COLS[i]}{r}"] = ""

# P2 actividades
ws["B16"] = "LL-2 · Drones agrarios · RGB (cartografía + regadío) + multiespectral (NDVI fitosanitario)"
mark(16, 2, "E1")  # Sep 2027

ws["B17"] = "LL-3 · Domo NEB · diseño digital + co-construcción (CAD → cálculo → estructura → cubierta → corcho)"
mark(17, 0, "E1")  # Jul 2027

ws["B18"] = "LL-4 · Living Lab digital · sensórica IoT + fotovoltaica + comunicaciones + volcado al SIG"
mark(18, 1, "E1")  # Ago 2027
mark(18, 2, "E1")  # Sep 2027

ws["B19"] = "LL-5 · Ciclo productivo estacional · 4 sesiones (otoño inicio · otoño cosecha · primavera inicio · primavera cosecha)"
mark(19, 2, "E1")  # Sep 2027
mark(19, 5, "E2")  # Dic 2027
mark(19, 7, "E3")  # Feb 2028
mark(19, 9, "E4")  # Abr 2028

ws["B20"] = "LL-6 · Jornada final NEB-Alcarria · resultados, AKIS y replicabilidad"
mark(20, 9, "E1")  # Abr 2028

# Limpiar casillas inexistentes (índices 10-11) en P2
for r in (16, 17, 18, 19, 20):
    for i in range(10, 12):
        ws[f"{MONTH_COLS[i]}{r}"] = ""

# --- G. Firma (A22 tras inserción) ---
ws["A22"] = "En Madrid, a 18 de mayo de 2026"

# --- H. Ampliar print area (era A1:P21, ahora hay 2 filas más) ---
ws.print_area = "'CRONOGRAMA'!$A$1:$P$23"

# --- I. Permitir que B5 (entidad) muestre 3 líneas (la cadena ocupa ~3 al
# wrappearse en B5:C5, anchura combinada ~54 chars).
ws.row_dimensions[5].height = 48.0

wb.save(SRC)
print(f"OK · escrito {SRC}")
